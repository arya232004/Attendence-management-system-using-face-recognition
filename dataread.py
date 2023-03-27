import openpyxl
dataframe = openpyxl.load_workbook("attendence.xlsx")
dataframe1 = dataframe.active
for row in range(1, dataframe1.max_row):
    li=[]
    stren=''
    for col in dataframe1.iter_cols(3, 3):
        stren=col[row].value
        for i in stren.split(","):
            li.append(float(i)) 
    print(li)      
        
    