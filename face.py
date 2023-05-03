from tkinter import messagebox
import cv2 as cv
import tkinter as tk
import numpy as np
import pandas 
import datetime
import time
import openpyxl
import openpyxl.utils
import face_recognition
  

root = tk.Tk()
c = tk.Canvas(root, bg="pink", height="200")
def check_fields(name_entry, roll_entry, submit_button, button_cam):
    if roll_entry.get() and name_entry.get():
        submit_button.config(state="normal")
        button_cam.config(state="normal")
    else:
        submit_button.config(state="disabled")
        button_cam.config(state="disabled")
    
def adduser(sheet_name):
    heading_font = ("Helvetica", 24, "bold")
    label_font = ("Helvetica", 12)
    entry_font = ("Helvetica", 12)
    
    print("in adduser",sheet_name)
    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True)

    name = tk.Label(frame, text="Enter name of student",font=heading_font, bg="#f2f2f2")
    name.pack(pady=20)

    name_entry = tk.Entry(frame,font=heading_font, bg="#f2f2f2")
    name_entry.pack(pady=5)
    name_entry.bind("<KeyRelease>", lambda event: check_fields(name_entry, roll_entry, submit_button, button_cam))

    roll = tk.Label(frame, text="Enter Roll no",font=heading_font, bg="#f2f2f2")
    roll.pack()

    roll_entry = tk.Entry(frame,font=heading_font, bg="#f2f2f2")
    roll_entry.pack()
    roll_entry.bind("<KeyRelease>", lambda event: check_fields(name_entry, roll_entry, submit_button, button_cam))

    button_cam = tk.Button(frame,font=label_font, bg="#b4d8e7",width=16 ,fg="#ffffff", text="Take Photo",command=lambda: takephoto(name_entry.get(), roll_entry.get(),sheet_name),height=1, state="disabled")
    button_cam.pack(pady=10)

    submit_button = tk.Button(frame, text="Submit", state="disabled",bg="#b4d8e7",width=16 ,fg="#ffffff")
    submit_button.pack(pady=10)

    check_fields(name_entry, roll_entry, submit_button, button_cam)

    button = tk.Button(frame, text="Close", command=frame.destroy,font=label_font)
    button.pack(pady=10)


def displaydata(frame,name,face_locations,roll):
     display=name+roll
     for (top, right, bottom, left) in face_locations:
            cv.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
            cv.putText(frame, display, (left, top), cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
            
            
def attendance(sheet_name):
    print("in attendence ",sheet_name)
    col_no = 0
    known_encodings = []
    name = "xddd"
    roll = ""
    dataframe = openpyxl.load_workbook("attendence.xlsx")
    dataframe1 = dataframe[sheet_name]
    
    col_names = [column[0].value for column in dataframe1.iter_cols(min_row=1, max_row=1)]
    print(col_names)
    today = datetime.date.today().strftime('%Y-%m-%d')
    print(today)
    
    if today in col_names:
        print("Already present")
        col_no = col_names.index(today) + 1
        print(col_no)    
    else:
        print("Not present")   
        col_no = dataframe1.max_column 
        print(col_no) 
        new_col_header = today
        dataframe1.insert_cols(col_no)
        dataframe1.cell(row=1, column=col_no).value = new_col_header

        if col_no == dataframe1.max_column:
            for row in range(2, dataframe1.max_row + 1):
                cell = dataframe1.cell(row=row, column=col_no)
                cell.value = '=COUNTIF(D' + str(row) + ':Z' + str(row) + ', "P")/COUNTA(D' + str(row) + ':Z' + str(row) + ')*100'
        dataframe.save("attendence.xlsx")
    
    for row in range(2, dataframe1.max_row + 1):
        li = []
        stren = ''
        for col in dataframe1.iter_cols(3, 3, min_row=row, max_row=row):
            stren = col[0].value
            for i in stren.split(","):
                li.append(float(i))  
        known_encodings.append(li)
    
    cam = cv.VideoCapture(0)
    while True:
        ret, frame = cam.read()
        rgb_frame = frame[:, :, ::-1]
        face_locations = face_recognition.face_locations(rgb_frame, model="hog")
        
        if len(face_locations) == 1:
            live_encoding = face_recognition.face_encodings(frame, face_locations)
            if live_encoding:
                results = face_recognition.compare_faces(known_encodings, live_encoding[0], tolerance=0.4)
                print(face_recognition.face_distance(known_encodings,live_encoding[0]))
                print(results)
                if True in results or name=="" or roll=="":
                    for i in range(len(results)):
                        if results[i]==True:
                            row = i + 2
                            break
                    print(row)        
                    print("Match Found")
                    name = dataframe1.cell(row, 2).value
                    print(name)
                    roll = dataframe1.cell(row, 1).value 
                    cell = dataframe1.cell(row, col_no)
                    cell.value = "P"
                    dataframe.save("attendence.xlsx")
                    displaydata(frame, name, face_locations, roll)
                else:
                    pass
                    print("Match Not Found") 
                    name = "un"
                    roll = "known"
                    print(name)
                    displaydata(frame, name, face_locations, roll)
            else:
                print("No face detected")
        elif len(face_locations) > 1:
            print("No of faces should be 1 only")
        else:
            print("No face detected")
        
        cv.imshow(sheet_name, frame)
        
        if cv.waitKey(1) & 0xFF == ord('q'):
            cam.release()
            cv.destroyAllWindows()   
            break

    
def takephoto(name,roll,sheet_name):
    workb = openpyxl.load_workbook("attendence.xlsx")
    worksheet = workb[sheet_name]
    last_row = worksheet.max_row
    face_encode_list=[]
    cam_port = 0
    cam = cv.VideoCapture(cam_port)
    time.sleep(2) 
    result, image = cam.read()
    print(result, image)
    if result:
        cv.imshow(name+roll+".jpg", image)
        cv.imwrite(name+roll+".jpg", image)
        
        imagefr = face_recognition.load_image_file(name+roll+".jpg")
        face_locations = face_recognition.face_locations(imagefr)
        print(len(face_locations))
        if(len(face_locations)==1):
           face_encodings = face_recognition.face_encodings(imagefr, face_locations)
           for i in range(len(face_encodings[0])):
            face_encode_list.append(face_encodings[0][i])  
           with open("Arya7.txt", "a") as file:
            for i in range(len(face_encodings[0])):
                file.write(str(face_encodings[0][i]))
                file.write("\n")
            row_to_insert = last_row + 1            
            worksheet.insert_rows(row_to_insert)    
            roll_no_cell = worksheet.cell(row=row_to_insert, column=1)
            name_cell = worksheet.cell(row=row_to_insert, column=2)
            face_encod=worksheet.cell(row=row_to_insert, column=3)
            name_cell.value = name
            roll_no_cell.value = roll 
            face_encod.value=','.join(map(str, face_encodings[0]))
            
            for row in worksheet.iter_cols(1,worksheet.max_column):
                #  print(row[0].value)
                if row[0].value == 'Attendence':
                        column_number = row[0].column
                        break
            cell = worksheet.cell(row=worksheet.max_row, column=column_number)
            cell.value ='=IFERROR(COUNTIF(D' + str(row_to_insert) + ':' + openpyxl.utils.cell.get_column_letter(worksheet.max_column) + str(row_to_insert) + ', "P")/COUNTIF(D' + str(row_to_insert) + ':' + openpyxl.utils.cell.get_column_letter(worksheet.max_column) + str(row_to_insert) + ', "*")*100, 100)'

            workb.save('attendence.xlsx')
            cv.waitKey(0)
        else:
            print("Error")
        cv.destroyWindow(name+roll+".jpg")
    else:
        print("No image detected. Please! try again")
    c.pack()
    root.mainloop()

def namedata(name,submit_button):
    if name.get():
        submit_button.config(state="normal")
    else:
        submit_button.config(state="disabled")    

def createnewhseet(name):
    print(name)
    workbook = openpyxl.load_workbook("attendence.xlsx")
    if(name in workbook.sheetnames):
        messagebox.showwarning("Sheet already exists")
    else:
      newsheet = workbook.create_sheet(title=name)
      newsheet['A1'] = "Roll no"
      newsheet['B1']="Name"
      
      newsheet['C1']="Face Encoding"
      newsheet['D1']="Attendence"
      workbook.save("attendence.xlsx")
      messagebox.showinfo("Sheet created successfully")
        
def addsheet():
    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True)    
    heading_font = ("Helvetica", 24, "bold")
    label_font = ("Helvetica", 12)
    entry_font = ("Helvetica", 12)
    name = tk.Label(frame, text="Enter name of sheet",font=heading_font, bg="#f2f2f2")
    name.pack(pady=20)
    # name.place(x=200, y=20)
    name_entry = tk.Entry(frame,font=heading_font, bg="#f2f2f2")
    name_entry.pack(pady=5)
    # name_entry.place(x=200, y=60)
    name_entry.bind("<KeyRelease>", lambda event: namedata(name_entry,submit_button))
    
    submit_button = tk.Button(frame, text="Submit",font=label_font, bg="#b4d8e7",width=14 ,fg="#ffffff" ,state="disabled", command=lambda: createnewhseet(name_entry.get()))
    submit_button.pack(pady=5)
    button = tk.Button(frame, text="Close", command=frame.destroy,font=label_font)
    button.pack(pady=5)
    
    
def listsheets(fun_name):
    heading_font = ("Helvetica", 24, "bold")
    label_font = ("Helvetica", 12)
    entry_font = ("Helvetica", 12)
    def printnow():
        selval=selected_option.get()
        print(selval)
        # print(fun_name)
        if(fun_name=="attendence"):
           attendance(selval) 
           frame.destroy()
        elif(fun_name=="adduser"):
           adduser(selval)
           frame.destroy() 
    workbook=openpyxl.load_workbook("attendence.xlsx")
    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True)
    
    sheets_name=workbook.sheetnames
    print(type(sheets_name))
    sheets_name.pop(0)
    selected_option = tk.StringVar(value=' ')
    for name in sheets_name:
        button = tk.Radiobutton(frame, text=name, variable=selected_option, value=name)
        button.pack(pady=10, anchor="center", fill="y")    
    button = tk.Button(frame, text="Select sheet", command=printnow)
    button.pack(pady=10, anchor="center", fill="y")  
    
    button = tk.Button(frame, text="Close", command=frame.destroy)
    button.pack(pady=10, anchor="center", fill="y")   
     
     
def authenticate(usernmae,password):
     if(usernmae=="admin" and password=="admin"):
        messagebox.showinfo("Welcome", "Welcome to the system")  
        home()
     else:
        messagebox.showerror("Error", "Invalid username or password")   
     

def home():

  root.title("Attendence")
  root.geometry("500x500")
  root.configure(bg="white")
  btn = tk.Button(root, text="Add User", fg="red", command=lambda: listsheets('adduser'), width=19, height=2)
  btn.place(x=180, y=40)

  btn = tk.Button(root, text="Attendence", fg="red", command=lambda: listsheets('attendence'), width=19, height=2)
  btn.place(x=180, y=150)

  btn=tk.Button(root,text="Create a new sheet",fg="red",command=addsheet,width=19,height=2)
  btn.place(x=180,y=260)
  root.mainloop()
  

def admin():
    root = tk.Tk()
    root.title("Login")
    root.geometry("400x300")
    root.configure(bg="#f2f2f2")
    
    heading_font = ("Helvetica", 24, "bold")
    label_font = ("Helvetica", 12)
    entry_font = ("Helvetica", 12)
    
    heading_label = tk.Label(root, text="Login", font=heading_font, bg="#f2f2f2")
    heading_label.pack(pady=20)

# Create the username label and entry field
    username_label = tk.Label(root, text="Username", font=label_font, bg="#f2f2f2")
    username_label.pack(pady=5)
    username_entry = tk.Entry(root, font=entry_font, bg="#ffffff")
    username_entry.pack(pady=5)

# Create the password label and entry field
    password_label = tk.Label(root, text="Password", font=label_font, bg="#f2f2f2")
    password_label.pack(pady=5)
    password_entry = tk.Entry(root, font=entry_font, bg="#ffffff", show="*")
    password_entry.pack(pady=5)

# Create the login button
    login_button = tk.Button(root, text="Login", font=label_font, bg="#4CAF50", fg="#ffffff",command=lambda: authenticate(username_entry.get(),password_entry.get()) )
    login_button.pack(pady=20)

#Run the main loop
    root.mainloop()
    
admin()   
